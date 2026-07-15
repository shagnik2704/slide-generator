import os
import re
import logging
from src.nodes.redesign.utils.schema import SharedAgentState, TutorialData, TutorialState, OldTutorial, UpdatedTutorial, GenerateTutorialRequest, GenerateTutorialResponse
from src.nodes.redesign.extract_links import fetch_links
from src.nodes.redesign.extraction import extract_tutorials
from src.nodes.redesign.updates import tech_intelligence_agent
from src.nodes.redesign.split import duration_split
from src.nodes.redesign.tabulate import form_final_table

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

import requests

def update_progress(task_id: str | None, webhook_url: str | None, status: str, progress: int, message: str, stage: str, url: str | None = None):
    if not task_id:
        return
    payload = {
        "task_id": task_id,
        "status": status,
        "progress": progress,
        "message": message,
        "stage": stage,
        "url": url
    }
    logger.info(f"Task {task_id} [{stage}] - Progress: {progress}% - {message}")
    if webhook_url:
        try:
            res = requests.post(webhook_url, json=payload, timeout=5)
            res.raise_for_status()
        except Exception as e:
            logger.warning(f"Failed to post to webhook {webhook_url}: {e}")

def run_pipeline(request: GenerateTutorialRequest, task_id: str | None = None, webhook_url: str | None = None) -> GenerateTutorialResponse:
    try:
        # Initial status
        update_progress(task_id, webhook_url, "processing", 5, "Initializing workspace...", "init")
        
        # Initialize TutorialData
        foss_name = request.foss_name
        language = request.language
        data = TutorialData(foss_name=foss_name, language=language, links=[])
        
        # Define output CSV path (temporary file for appending raw data)
        safe_foss = re.sub(r'[^a-zA-Z0-9_\-]', '_', foss_name)
        os.makedirs("results", exist_ok=True)
        output_csv_path = os.path.join("results", f"temp_{safe_foss}_{language}.csv")
        
        # Clean output CSV if it already exists
        if os.path.exists(output_csv_path):
            try:
                os.remove(output_csv_path)
            except OSError:
                pass
                
        # Initialize SharedAgentState
        state = SharedAgentState(
            data=data,
            output_csv_path=output_csv_path,
            tutorial=None
        )
        
        # 1. Fetch links
        update_progress(task_id, webhook_url, "processing", 10, "Fetching tutorial links from Spoken Tutorial website...", "fetch_links")
        state.data = fetch_links(state.data)
        
        # Check if any tutorials were found
        if not state.data.links:
            raise ValueError(
                f"No tutorials found for '{foss_name}' in '{language}'. "
                f"This FOSS might not be available in the selected language."
            )
        
        num_links = len(state.data.links)
        update_progress(task_id, webhook_url, "processing", 20, f"Found {num_links} tutorials. Starting content redesign...", "analysis_start")
        
        # 2. Iterate over links
        new_t_counter = 1
        for idx, link in enumerate(state.data.links, start=1):
            # Calculate progress bounds for this iteration
            tut_start_progress = 20 + int((idx - 1) * 70 / num_links)
            tut_end_progress = 20 + int(idx * 70 / num_links)
            step_delta = max(1, (tut_end_progress - tut_start_progress) // 4)
            
            # Initialize TutorialState for this loop iteration
            state.tutorial = TutorialState(
                tutorial_name=link.name,
                tutorial_link=link.url,
                old_tutorial=OldTutorial(),
                updated_tutorial=UpdatedTutorial(),
                splited_tutorial=[]
            )
            
            # Run stages
            logger.info(f"Running extraction stage for {state.tutorial.tutorial_name}")
            update_progress(
                task_id, webhook_url, "processing", 
                tut_start_progress, 
                f"[{idx}/{num_links}] Extracting contents from: {state.tutorial.tutorial_name}", 
                "extraction"
            )
            state.tutorial = extract_tutorials(state.tutorial)
            
            logger.info(f"Running tech intelligence stage for {state.tutorial.tutorial_name}")
            update_progress(
                task_id, webhook_url, "processing", 
                tut_start_progress + step_delta, 
                f"[{idx}/{num_links}] Running tech intelligence agent for: {state.tutorial.tutorial_name}", 
                "tech_intelligence"
            )
            state.tutorial = tech_intelligence_agent(state.tutorial)
            
            logger.info(f"Running duration split stage for {state.tutorial.tutorial_name}")
            update_progress(
                task_id, webhook_url, "processing", 
                tut_start_progress + 2 * step_delta, 
                f"[{idx}/{num_links}] Splitting tutorial: {state.tutorial.tutorial_name}", 
                "duration_split"
            )
            state.tutorial = duration_split(state.tutorial)
            
            logger.info(f"Running tabulate stage for {state.tutorial.tutorial_name}")
            update_progress(
                task_id, webhook_url, "processing", 
                tut_start_progress + 3 * step_delta, 
                f"[{idx}/{num_links}] Saving table results for: {state.tutorial.tutorial_name}", 
                "tabulation"
            )
            state.tutorial, new_t_counter = form_final_table(
                state.tutorial, 
                output_csv_path=state.output_csv_path, 
                tutorial_index=idx,
                start_new_t_index=new_t_counter
            )

        # 3. Convert CSV to XLSX, calculate total durations, update row 2
        update_progress(task_id, webhook_url, "processing", 90, "Converting output to Excel format...", "export")
        
        output_xlsx_path = os.path.join("results", f"{safe_foss}_{language}_final.xlsx")
        
        import csv
        rows = []
        if os.path.exists(output_csv_path):
            with open(output_csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)

        import openpyxl
        template_xlsx_path = os.path.join("src", "nodes", "redesign", "utils", "VC-Template.xlsx")
        
        if os.path.exists(template_xlsx_path) and len(rows) > 1:
            wb = openpyxl.load_workbook(template_xlsx_path)
            ws = wb.active
            
            def parse_duration_to_day_fraction(val):
                if not val:
                    return 0.0
                parts = str(val).split(':')
                try:
                    if len(parts) == 3:
                        secs = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                    elif len(parts) == 2:
                        secs = int(parts[0]) * 60 + int(parts[1])
                    else:
                        secs = int(parts[0])
                    return secs / 86400.0
                except ValueError:
                    return 0.0

            # rows[0] is header. Data starts from index 1.
            data_rows = rows[1:]

            # Append redesign data rows starting from row 4
            start_row = 4
            for idx, row in enumerate(data_rows):
                current_row = start_row + idx
                for col_idx, val in enumerate(row):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    if col_idx == 3 or col_idx == 8:  # Columns D and I
                        cell.value = parse_duration_to_day_fraction(val)
                        cell.number_format = 'hh:mm:ss'
                    else:
                        cell.value = val

            # Write SUM formula in Row 2 columns 4 (D2) and 9 (I2)
            last_row = start_row + len(data_rows) - 1
            
            cell_d2 = ws.cell(row=2, column=4)
            cell_d2.value = f"=SUM(D4:D{last_row})"
            cell_d2.number_format = '[h]:mm:ss'
            
            cell_i2 = ws.cell(row=2, column=9)
            cell_i2.value = f"=SUM(I4:I{last_row})"
            cell_i2.number_format = '[h]:mm:ss'

            wb.save(output_xlsx_path)
            logger.info(f"Excel file successfully generated at {output_xlsx_path} using styled template.")
            
            try:
                os.remove(output_csv_path)
            except OSError:
                pass
        else:
            # Fallback to writing standard Excel sheet if template is missing
            import pandas as pd
            if len(rows) > 0:
                df_out = pd.DataFrame(rows)
                df_out.to_excel(output_xlsx_path, index=False, header=False)
            logger.warning(f"Template not found at {template_xlsx_path} or no data generated. Default Excel written.")
            
            try:
                os.remove(output_csv_path)
            except OSError:
                pass
        
        output_filename = f"{safe_foss}_{language}_final.xlsx"
        update_progress(task_id, webhook_url, "completed", 100, "Tutorial redesign completed successfully!", "completed", url=output_filename)
        return GenerateTutorialResponse(status="success", url=output_filename)
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Error in redesign pipeline: {error_msg}", exc_info=True)
        update_progress(task_id, webhook_url, "failed", 100, f"Error: {error_msg}", "failed")
        return GenerateTutorialResponse(status="error", url=error_msg)

